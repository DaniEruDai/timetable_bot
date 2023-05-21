import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from calendar import monthrange
from dateutil.relativedelta import relativedelta

def get_name_of_month(__date : str) -> str:
  return {1: f'Январь ',
    2: f'Феварль ',
    3: f'Март ',
    4: f'Апрель ',
    5: f'Май ',
    6: f'Июнь ',
    7: f'Июль ',
    8: f'Август ' ,
    9: f'Сентябрь ' ,
    10: f'Окятбрь ',
    11: f'Ноябрь ',
    12: f'Декабрь '}[int( datetime.strptime(__date,'%m.%Y').strftime('%m') )] + datetime.strptime(__date,'%m.%Y').strftime('%Y')

def get_year(__year : str) -> list :
  start = datetime.strptime(f'0109{__year}', "%d%m%Y").date()
  dates = [start.strftime('%m.%Y')]+ [(start + relativedelta(months=date)).strftime('%m.%Y') for date in range(1, 10)]
  return dates

class Style:
  
  def __init__(self,__WorkBook : openpyxl.Workbook):
    self.wb = __WorkBook
    self.bad = self.__bad()
    self.normal = self.__normal()
    self.neutral = self.__neutral()
    
  def __neutral (self):
    ns = NamedStyle(name='NEUTRAL')
    ns.fill = PatternFill(fgColor='fcfcee', fill_type='solid')
    ns.font = Font(color='1f0800', name='Bahnscrift', bold=False)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    self.wb.add_named_style(ns)
    return ns.name
  
  def __bad(self):
    ns = NamedStyle(name='BAD')
    ns.fill = PatternFill(fgColor='ff8e7a', fill_type='solid')
    ns.font = Font(color='1f0800', name='Bahnscrift', bold=True)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    self.wb.add_named_style(ns)
    return ns.name

  def __normal(self):
    ns = NamedStyle(name='NORMAL')
    ns.fill = PatternFill(fgColor='ffd500', fill_type='solid')
    ns.font = Font(color='1f0800', name='Bahnscrift', bold=True)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    self.wb.add_named_style(ns)
    return ns.name

class Excel:
  
  def __init__(self,__all_marks : list, __year : str) -> None:
    self.__year = __year
    self.MARKS = __all_marks
    self.__wb = openpyxl.Workbook()
    self.__sheet = self.__wb['Sheet']
    self.__matrix = self.__get_matrix()

  def __get_matrix(self):
    # Фильтрация оценок по заданному году
    self.MARKS = list(filter(lambda x : True if x.date.strftime('%m.%Y') in get_year(self.__year) else False ,self.MARKS))
    
    # Сортировка оценок по дате
    self.MARKS.sort(key = lambda x : x.date)
    
    matrix = []
    
    start = datetime.strptime(f'0109{self.__year}', "%d%m%Y").date()
    MonthYear = [start.strftime('%m.%Y')]+ [(start + relativedelta(months=date)).strftime('%m.%Y') for date in range(1, 10)]
    dates = []
    for _my in MonthYear:
      months = []
      days_in_month = monthrange(int(_my[3:]), int(_my[:2]))[1]
      for day in range(1,days_in_month+1):
        if len(str(day)) == 1:
          day = f'0{day}'
        months.append(f'{day}.{_my}')  
      dates.append(months)
    
      
    
    for m_Y,month_dates in zip(get_year(self.__year),dates):
      
      lessons = {i.lesson for i in self.MARKS if i.date.strftime('%m.%Y') == m_Y}
      lessons = list(lessons)
      lessons.sort()
      if not lessons:
        continue
      
      mark_of_month = [m for m in self.MARKS if m.date.strftime('%m.%Y') == m_Y]
      
      #HEADER - 'Месяц + год и числа месяца'
      header = [get_name_of_month(m_Y)] + [i for i in range(1,len(month_dates)+1)]
      while len(header) != 32:
          header.append('')
      header.append('Ср.балл')
      matrix.append(header)
      
      
      for lesson in lessons:
        mark_of_lessons = [m for m in mark_of_month if m.lesson == lesson]
        
        row = ['' for i in month_dates]
        
        #Создание 'месяца' с оценками
        for x in mark_of_lessons:
          str_date = x.date.strftime('%d.%m.%Y')
          index = month_dates.index(str_date)
          row[index] = x.mark
        
        #ROW - 'Предмет и его оценки'
        
        while len(row) != 31:
          row.append('')
        
        #Нахождение среднего балла
        digits = [x for x in row if isinstance(x,int)]
        
        average = sum(digits)/len(digits)
        
        average = round(average,2)
        if str(average).endswith('0'):
          average = round(average)
        row.append(average)
  
        matrix.append([lesson] + row)
        
      matrix.append([''])
      matrix.append([''])
        
    return matrix

  def __recolor(self,value : str | int , style_name : str):
    normal = [(index+1, row.index(value)+1) for index, row in enumerate(self.__matrix) if value in row]
    for row, column in normal:
     self.__sheet.cell(column=column, row=row).style = style_name
   
  def __fill_table(self):
  
    style = Style(self.__wb)
    
    #Заполнение Excel-файла
    for row in self.__matrix:
      self.__sheet.append(row)
    
    #Изменение названия листа
    self.__sheet.title = 'Оценки'
    
    #Изменение ширины столбцов
    for i in range(2, self.__sheet.max_column+1):
      self.__sheet.column_dimensions[f'{get_column_letter(i-1)}'].width = 5
    self.__sheet.column_dimensions['A'].width = 100
    
    #Изменение положения текста внутри ячеек
    for row in self.__sheet[f'A1:AO{self.__sheet.max_row}']:
      for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    self.__recolor(3,style.normal)
    self.__recolor(2,style.bad)
    
    for value in get_year(self.__year):
      value = get_name_of_month(value)
      indices = [(index+1, row.index(value)+1) for index, row in enumerate(self.__matrix) if value in row]
      for index in indices:
        for cell in self.__sheet[f"{index[0]}:{index[0]}"]:
          cell.style = style.neutral

  def save(self,__file : str | bytes) -> None:
    self.__fill_table()
    self.__wb.save(__file)


